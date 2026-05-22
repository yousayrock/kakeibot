#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
スケジュール管理モジュール
家系Bot用 - 予定の追加・修正・削除・Excel出力・繰り返し展開
"""

import json
import shutil
import logging
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger("kakeibo")

# 曜日マッピング
WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]
RECURRENCE_LABELS = {
    "none":    "なし",
    "daily":   "毎日",
    "weekly":  "毎週",
    "monthly": "毎月",
}


# ────────────────────────────────────────────
# ファイル操作
# ────────────────────────────────────────────

def get_schedule_dir(kakeibo_dir: Path, year: int, month: int) -> Path:
    d = kakeibo_dir / str(year) / f"{month:02d}_{month}月"
    d.mkdir(parents=True, exist_ok=True)
    return d


def load_schedules(kakeibo_dir: Path, year: int, month: int) -> list:
    p = get_schedule_dir(kakeibo_dir, year, month) / "schedule.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return []


def save_schedules(kakeibo_dir: Path, year: int, month: int, schedules: list):
    p = get_schedule_dir(kakeibo_dir, year, month) / "schedule.json"
    if p.exists():
        shutil.copy2(p, p.with_suffix(".bak"))
    p.write_text(json.dumps(schedules, ensure_ascii=False, indent=2), encoding="utf-8")


def make_event_id(date_str: str, title: str) -> str:
    ts   = datetime.now().strftime("%H%M%S%f")[:8]
    safe = "".join(c for c in title if c.isalnum() or '\u3000' <= c <= '\u9fff')[:10]
    return f"{date_str.replace('-', '')}_{ts}_{safe}"


def add_schedule(kakeibo_dir: Path, schedule: dict) -> dict:
    """予定を追加してJSONとExcelを更新"""
    year, month = schedule["year"], schedule["month"]
    schedules = load_schedules(kakeibo_dir, year, month)
    schedules.append(schedule)
    save_schedules(kakeibo_dir, year, month, schedules)
    update_excel_schedule(kakeibo_dir, year, month, schedules)

    # 繰り返しの展開（翌月分まで）
    if schedule.get("recurrence", "none") != "none":
        expand_recurring(kakeibo_dir, schedule, months_ahead=3)

    logger.info(f"📅 予定追加: {schedule['title']} {schedule['date']}")
    return schedule


def update_schedule_by_id(kakeibo_dir: Path, year: int, month: int,
                           event_id: str, updates: dict) -> bool:
    """IDで予定を検索して更新"""
    schedules = load_schedules(kakeibo_dir, year, month)
    for i, s in enumerate(schedules):
        if s.get("id") == event_id:
            schedules[i].update(updates)
            save_schedules(kakeibo_dir, year, month, schedules)
            update_excel_schedule(kakeibo_dir, year, month, schedules)
            return True
    return False


def delete_schedule_by_id(kakeibo_dir: Path, year: int, month: int,
                           event_id: str) -> dict | None:
    """IDで予定を削除"""
    schedules = load_schedules(kakeibo_dir, year, month)
    for i, s in enumerate(schedules):
        if s.get("id") == event_id:
            deleted = schedules.pop(i)
            save_schedules(kakeibo_dir, year, month, schedules)
            update_excel_schedule(kakeibo_dir, year, month, schedules)
            return deleted
    return None


def expand_recurring(kakeibo_dir: Path, base: dict, months_ahead: int = 3):
    """繰り返し予定を翌月以降に展開"""
    recurrence = base.get("recurrence", "none")
    if recurrence == "none":
        return

    base_date = datetime.strptime(base["date"], "%Y-%m-%d")
    now       = datetime.now()

    # 展開する日付を生成
    dates_to_add = []
    if recurrence == "daily":
        current = base_date + timedelta(days=1)
        end     = now + timedelta(days=months_ahead * 30)
        while current <= end:
            dates_to_add.append(current)
            current += timedelta(days=1)

    elif recurrence == "weekly":
        current = base_date + timedelta(weeks=1)
        end     = now + timedelta(days=months_ahead * 30)
        while current <= end:
            dates_to_add.append(current)
            current += timedelta(weeks=1)

    elif recurrence == "monthly":
        for m in range(1, months_ahead + 1):
            month_offset = (base_date.month - 1 + m) % 12 + 1
            year_offset  = base_date.year + (base_date.month - 1 + m) // 12
            try:
                dates_to_add.append(base_date.replace(year=year_offset, month=month_offset))
            except ValueError:
                pass  # 月末超え（例：1/31→2/31）はスキップ

    for dt in dates_to_add:
        year, month = dt.year, dt.month
        existing = load_schedules(kakeibo_dir, year, month)
        # 同じイベントIDベースの繰り返しが既に存在しないか確認
        base_id = base["id"]
        already = any(s.get("recurrence_parent_id") == base_id
                      and s.get("date") == dt.strftime("%Y-%m-%d")
                      for s in existing)
        if already:
            continue

        new_event = base.copy()
        new_event["date"]               = dt.strftime("%Y-%m-%d")
        new_event["year"]               = dt.year
        new_event["month"]              = dt.month
        new_event["id"]                 = make_event_id(dt.strftime("%Y-%m-%d"), base["title"])
        new_event["recurrence_parent_id"] = base_id
        new_event["google_event_id"]    = None  # 繰り返しはGoogle側で管理

        existing.append(new_event)
        save_schedules(kakeibo_dir, year, month, existing)
        update_excel_schedule(kakeibo_dir, year, month, existing)


# ────────────────────────────────────────────
# Excel出力
# ────────────────────────────────────────────

def update_excel_schedule(kakeibo_dir: Path, year: int, month: int, schedules: list):
    """スケジュールExcelを更新"""
    path = get_schedule_dir(kakeibo_dir, year, month) / "schedule.xlsx"
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = f"{year}年{month}月スケジュール"

    # スタイル定義
    hfill   = PatternFill("solid", fgColor="1F4E79")
    shopfill= PatternFill("solid", fgColor="D9E1F2")
    perfill = PatternFill("solid", fgColor="E2EFDA")
    hfont   = Font(bold=True, color="FFFFFF")
    border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ヘッダー
    headers = ["日付", "曜日", "開始", "終了", "タイトル", "場所", "カレンダー", "メモ", "繰り返し"]
    widths  = [12, 6, 8, 8, 25, 20, 10, 25, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = border
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col)].width = w

    # データ
    sorted_schedules = sorted(schedules, key=lambda x: (x.get("date", ""), x.get("start_time", "")))
    for row_idx, s in enumerate(sorted_schedules, 2):
        try:
            dt      = datetime.strptime(s["date"], "%Y-%m-%d")
            weekday = WEEKDAY_JP[dt.weekday()]
        except Exception:
            dt      = None
            weekday = ""

        cal_name = "お店" if s.get("calendar") == "shop" else "個人"
        fill     = shopfill if s.get("calendar") == "shop" else perfill
        rec_label= RECURRENCE_LABELS.get(s.get("recurrence", "none"), "なし")

        values = [
            s.get("date", ""), weekday,
            s.get("start_time", ""), s.get("end_time", ""),
            s.get("title", ""), s.get("location", ""),
            cal_name, s.get("memo", ""), rec_label,
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.fill = fill; c.border = border
            c.alignment = Alignment(horizontal="center" if col <= 4 else "left")

    wb.save(path)
    logger.info(f"📊 schedule.xlsx 更新: {year}/{month}")


# ────────────────────────────────────────────
# 表示用フォーマット
# ────────────────────────────────────────────

def format_schedule_confirm(s: dict) -> str:
    """確認メッセージ用フォーマット"""
    try:
        dt      = datetime.strptime(s["date"], "%Y-%m-%d")
        weekday = WEEKDAY_JP[dt.weekday()]
        date_str= f"{dt.month}/{dt.day}（{weekday}）"
    except Exception:
        date_str = s.get("date", "")

    start = s.get("start_time", "未定")
    end   = s.get("end_time", "")
    time_str = f"{start} 〜 {end}" if end else start

    cal_name = "お店" if s.get("calendar") == "shop" else "個人"
    rec_label= RECURRENCE_LABELS.get(s.get("recurrence", "none"), "なし")

    lines = [
        f"タイトル　：{s.get('title', '')}",
        f"カレンダー：{cal_name}",
        f"日付　　　：{date_str}",
        f"時間　　　：{time_str}",
    ]
    if s.get("location"):
        lines.append(f"場所　　　：{s['location']}")
    if s.get("memo"):
        lines.append(f"メモ　　　：{s['memo']}")
    lines.append(f"繰り返し　：{rec_label}")
    lines.append(f"リマインダー：{s.get('reminder_minutes', 30)}分前")

    return "```\n" + "\n".join(lines) + "\n```"


def format_schedule_line(s: dict) -> str:
    """一覧表示用1行フォーマット"""
    try:
        dt      = datetime.strptime(s["date"], "%Y-%m-%d")
        weekday = WEEKDAY_JP[dt.weekday()]
        date_str= f"{dt.month}/{dt.day}（{weekday}）"
    except Exception:
        date_str = s.get("date", "")

    start    = s.get("start_time", "")
    cal_name = "🏪" if s.get("calendar") == "shop" else "👤"
    title    = s.get("title", "（無題）")
    time_part= f" {start}" if start else ""
    return f"{cal_name} {date_str}{time_part}　{title}"


def parse_time_str(text: str) -> str | None:
    """「10時」「10:00」「10時30分」などを「HH:MM」に変換"""
    text = text.strip().replace("：", ":")
    # 「なし」「スキップ」→ None
    if text in ("なし", "スキップ", "skip", "none", ""):
        return None
    # HH:MM 形式
    m = __import__("re").match(r'^(\d{1,2}):(\d{2})$', text)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    # 「10時」「10時30分」形式
    m = __import__("re").match(r'^(\d{1,2})時(?:(\d{1,2})分)?$', text)
    if m:
        h = int(m.group(1))
        mi = int(m.group(2)) if m.group(2) else 0
        return f"{h:02d}:{mi:02d}"
    # 数字のみ（「10」→「10:00」）
    m = __import__("re").match(r'^(\d{1,2})$', text)
    if m:
        return f"{int(m.group(1)):02d}:00"
    return None
