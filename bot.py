#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
家計簿Bot - レシート自動読み取り・家計簿管理
"""

import os
import re
import asyncio
import json
import shutil
import base64
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from datetime import datetime

import discord
from discord.ext import commands, tasks
from dotenv import load_dotenv
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import aiohttp

try:
    import nisshi  # 航海日誌モジュール
except ImportError:
    nisshi = None  # 未インストールでもbotは起動する

try:
    import drive_sync  # Google Driveデータ永続化
except ImportError:
    drive_sync = None

load_dotenv()

DISCORD_TOKEN      = os.getenv("DISCORD_TOKEN")
ANTHROPIC_KEY      = os.getenv("ANTHROPIC_API_KEY")
KAKEIBO_DIR        = Path(os.getenv("KAKEIBO_DIR", "data"))
OUTPUT_DIR         = Path(os.getenv("OUTPUT_DIR", "output"))
KAKEIBO_CHANNEL_ID = int(os.getenv("KAKEIBO_CHANNEL_ID", "0")) or None

KAKEIBO_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

CONFIG_PATH = KAKEIBO_DIR / "config.json"
STATE_PATH  = KAKEIBO_DIR / "state.json"   # 再起動時の状態復元用

ai_client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

CATEGORIES = [
    "食費", "交通費", "通信費", "消耗品費", "仕事経費",
    "光熱費", "医療費", "娯楽費", "外食費", "衣服費", "日用品", "その他",
]

# ────────────────────────────────────────────
# ログ設定
# ────────────────────────────────────────────

logger = logging.getLogger("kakeibo")
logger.setLevel(logging.INFO)

_handler = RotatingFileHandler("kakeibo.log", maxBytes=200_000, backupCount=2, encoding="utf-8")
_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S"))
logger.addHandler(_handler)

_error_handler = logging.FileHandler("kakeibo_error.log", encoding="utf-8")
_error_handler.setLevel(logging.WARNING)
_error_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S"))
logger.addHandler(_error_handler)

_console = logging.StreamHandler()
_console.setFormatter(logging.Formatter("%(asctime)s %(message)s", "%H:%M:%S"))
logger.addHandler(_console)

# ────────────────────────────────────────────
# Discord Bot
# ────────────────────────────────────────────

intents = discord.Intents.default()
intents.message_content = True
intents.messages        = True
intents.guilds          = True
bot = commands.Bot(command_prefix="!", intents=intents)

# 状態管理
pending_date:        dict[int, dict] = {}  # 日付確認待ち
pending_delete:      dict[int, dict] = {}  # 削除確認待ち
pending_confirm:     dict[int, dict] = {}  # 手入力確認待ち
pending_manual:      dict[int, dict] = {}  # 手入力受付中
pending_reply_edit:  dict[int, dict] = {}  # リプライ修正確認待ち
last_record_context: dict[int, dict] = {}  # 直近の記録コンテキスト（MAX_CTX_ENTRIES件まで）
onboarding_state:    dict[int, dict] = {}  # 初回ヒアリング進行中
pending_duplicate:   dict[int, dict] = {}  # 重複確認待ち
accounting_mode:     dict[int, bool]  = {}  # 経理モード中のユーザー


# ────────────────────────────────────────────
# 設定（config.json）
# ────────────────────────────────────────────

def load_config() -> dict:
    if CONFIG_PATH.exists():
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    return {}

def save_config(config: dict):
    CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

def load_state() -> dict:
    """pending状態をファイルから復元する。"""
    if STATE_PATH.exists():
        try:
            return json.loads(STATE_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_state(state: dict):
    """pending状態をファイルに保存する。"""
    try:
        STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, default=str), encoding="utf-8")
    except Exception as e:
        logger.warning(f"state保存失敗: {e}")

def flush_state(user_id: int):
    """ユーザーの状態をstateファイルから削除する。"""
    st = load_state()
    st.pop(str(user_id), None)
    save_state(st)

def persist_state(user_id: int, key: str, value: dict):
    """ユーザーの特定状態を永続化する。"""
    st = load_state()
    if str(user_id) not in st:
        st[str(user_id)] = {}
    st[str(user_id)][key] = value
    save_state(st)


# ────────────────────────────────────────────
# 定期費用・予算
# ────────────────────────────────────────────

def get_recurring() -> list:
    return load_config().get("recurring", [])

def get_budgets() -> dict:
    return load_config().get("budgets", {})

def set_budgets(budgets: dict):
    cfg = load_config(); cfg["budgets"] = budgets; save_config(cfg)

def set_recurring(items: list):
    cfg = load_config(); cfg["recurring"] = items; save_config(cfg)


# ────────────────────────────────────────────
# 過去記録の検索
# ────────────────────────────────────────────

def search_records(user_id: int, keyword: str = "", min_amount: int = 0,
                   max_amount: int = 0, category: str = "") -> list:
    """全月のレコードを横断検索。"""
    results = []
    now = datetime.now()
    for year in range(now.year - 1, now.year + 1):
        for month in range(1, 13):
            for r in load_records(year, month, user_id):
                if keyword and keyword not in r.get("name","") + r.get("note",""):
                    continue
                if min_amount and r.get("amount", 0) < min_amount:
                    continue
                if max_amount and r.get("amount", 0) > max_amount:
                    continue
                if category and r.get("category","") != category:
                    continue
                results.append(r)
    return sorted(results, key=lambda x: x.get("date",""), reverse=True)


# ────────────────────────────────────────────
# 経理AIアドバイザー（v1.2）
# ────────────────────────────────────────────

ACCOUNTING_SYSTEM = """あなたは日本の税務・経理の専門AIアシスタントです。
個人事業主・フリーランス向けに、確定申告・経費処理・消費税・インボイス制度などについて
正確でわかりやすいアドバイスを日本語で提供してください。

回答の最初に必ず以下の形式でJSONを1行だけ出力してください：
{"complexity": "simple" または "complex", "needs_specialist": true または false}

complexityの判断基準：
- simple：一般的な経費判断、カテゴリ分類、簡単な節税tips
- complex：税務調査対応、減価償却の細かい計算、インボイス登録の判断、
           複雑な按分計算、法人化の検討、特殊な業種の経費処理

JSONの後に改行して、実際の回答を書いてください。"""

async def ask_accounting(question: str, user_id: int) -> tuple[str, str]:
    """
    経理AIに質問する。
    Step1: Haikuで回答 + 複雑度評価
    Step2: complexなら Sonnet でより詳細な回答
    Returns: (回答テキスト, 使用モデル名)
    """
    profile = get_business_profile()
    context = f"【事業プロフィール】{profile}\n\n" if profile else ""

    # Step1: Haiku で高速回答
    try:
        r1 = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=800,
            system=ACCOUNTING_SYSTEM,
            messages=[{"role": "user", "content": f"{context}質問：{question}"}],
        )
        full = r1.content[0].text.strip()
    except Exception as e:
        logger.error(f"経理AI(Haiku)エラー: {e}")
        return "❌ AIへの接続に失敗しました。しばらく待ってから再試行してください。", "error"

    # JSON部分を抽出
    import json as _json
    complexity = "simple"
    try:
        first_line = full.split("\n")[0].strip()
        meta = _json.loads(first_line)
        complexity = meta.get("complexity", "simple")
        answer = "\n".join(full.split("\n")[1:]).strip()
    except Exception:
        answer = full

    # Step2: complex なら Sonnet で再回答
    if complexity == "complex":
        try:
            r2 = ai_client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=1500,
                system=ACCOUNTING_SYSTEM.replace(
                    'JSONの後に改行して、実際の回答を書いてください。',
                    'JSONの後に改行して、専門的かつ詳細な回答を書いてください。具体的な数字や法的根拠があれば添えてください。'
                ),
                messages=[{"role": "user", "content": f"{context}質問：{question}"}],
            )
            full2 = r2.content[0].text.strip()
            try:
                answer = "\n".join(full2.split("\n")[1:]).strip()
            except Exception:
                answer = full2
            return answer, "Sonnet（専門）"
        except Exception as e:
            logger.error(f"経理AI(Sonnet)エラー: {e}")
            # Haikuの回答で代替
            return answer, "Haiku（簡易）"

    return answer, "Haiku（簡易）"


def get_business_profile() -> str:
    """analyze_receipt に渡す事業プロフィール文を返す。未設定は空文字。"""
    return load_config().get("business_profile", "")

def is_onboarding_done() -> bool:
    # Railway環境では環境変数で管理（ファイルは再起動で消える）
    if os.environ.get("ONBOARDING_DONE", "").lower() == "true":
        return True
    return load_config().get("onboarding_done", False)


# ────────────────────────────────────────────
# ユーティリティ（マルチユーザー対応）
# ────────────────────────────────────────────

def get_month_dir(year: int, month: int, user_id: int = 0) -> Path:
    d = KAKEIBO_DIR / str(user_id) / str(year) / f"{month:02d}_{month}月"
    d.mkdir(parents=True, exist_ok=True)
    (d / "receipts").mkdir(exist_ok=True)
    return d


def get_excel_path(year: int, month: int, user_id: int = 0) -> Path:
    return get_month_dir(year, month, user_id) / "kakeibo.xlsx"


def load_records(year: int, month: int, user_id: int = 0) -> list:
    p = get_month_dir(year, month, user_id) / "records.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return []


def save_records(year: int, month: int, records: list, user_id: int = 0):
    p   = get_month_dir(year, month, user_id) / "records.json"
    tmp = p.with_suffix(".tmp")
    try:
        if p.exists():
            shutil.copy2(p, p.with_suffix(".bak"))
        tmp.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(p)
        # Driveにバックアップ
        if drive_sync:
            drive_sync.upload(p, f"{user_id}/{year}/{month:02d}/records")
    except Exception as e:
        logger.error(f"records.json 書き込み失敗: {e}")
        if tmp.exists(): tmp.unlink()
        raise


def make_record_id(date_str: str, filename: str) -> str:
    ts   = datetime.now().strftime("%H%M%S%f")[:10]
    safe = re.sub(r'[^\w.]', '_', filename)
    return f"{date_str.replace('-', '')}_{ts}_{safe}"


def sanitize_filename(name: str) -> str:
    """ファイル名に使えない文字を除去する。"""
    import re as _re
    name = _re.sub(r'[\\/:*?"<>|\r\n\t]', '_', name)
    name = name.strip('. ')
    return name[:50] or "receipt"


def make_receipt_filename(date_str: str, store_name: str, original: str,
                           receipt_time: str | None, sent_at: datetime) -> str:
    date_compact = date_str.replace("-", "")
    time_compact = receipt_time.replace(":", "")[:4] if receipt_time else sent_at.strftime("%H%M%S")
    ext          = original.rsplit(".", 1)[-1].lower() if "." in original else "jpg"
    safe_store   = sanitize_filename(store_name)[:20]
    return f"{date_compact}_{time_compact}_{safe_store}.{ext}"


def add_record(record: dict):
    year, month, uid = record["year"], record["month"], record.get("user_id", 0)
    records = load_records(year, month, uid)
    records.append(record)
    save_records(year, month, records, uid)
    update_excel(year, month, records, uid)
    # 予算アラートチェック（非同期チャンネル通知はイベントループ経由）
    if record.get("type") == "支出":
        _check_budget_sync(record, records)

def _check_budget_sync(new_record: dict, all_records: list):
    """予算チェック。超過・80%到達時にフラグをconfig保存（チャンネル通知はon_message側で拾う）。"""
    budgets = get_budgets()
    cat = new_record.get("category", "")
    if cat not in budgets:
        return
    limit = budgets[cat]
    spent = sum(r.get("amount", 0) for r in all_records
                if r.get("type") == "支出" and r.get("category") == cat)
    ratio = spent / limit if limit else 0
    cfg = load_config()
    alerts = cfg.setdefault("budget_alerts", {})
    key = f"{cat}_{new_record.get('year',0)}_{new_record.get('month',0)}"
    if ratio >= 1.0 and alerts.get(key) != "over":
        alerts[key] = "over"
        cfg["budget_alerts"] = alerts
        save_config(cfg)
        cfg["_pending_alert"] = f"🚨 **{cat}**の予算オーバー！ ¥{spent:,} / ¥{limit:,}（{ratio*100:.0f}%）"
        save_config(cfg)
    elif ratio >= 0.8 and alerts.get(key) not in ("over", "warn"):
        alerts[key] = "warn"
        cfg["budget_alerts"] = alerts
        cfg["_pending_alert"] = f"⚠️ **{cat}**が予算の{ratio*100:.0f}%に到達しました（¥{spent:,} / ¥{limit:,}）"
        save_config(cfg)


MAX_CTX_ENTRIES = 500  # last_record_context の上限

def set_last_context(user_id: int, record: dict):
    # メモリ上限チェック
    if len(last_record_context) >= MAX_CTX_ENTRIES:
        oldest = next(iter(last_record_context))
        last_record_context.pop(oldest, None)
    records = load_records(record["year"], record["month"], user_id)
    idx = next((i for i, r in enumerate(records) if r.get("id") == record.get("id")), len(records) - 1)
    last_record_context[user_id] = {
        "record": record,
        "year":   record["year"],
        "month":  record["month"],
        "index":  idx,
    }


def find_record_by_id(year: int, month: int, record_id: str, user_id: int = 0) -> tuple[int, dict] | tuple[None, None]:
    for i, r in enumerate(load_records(year, month, user_id)):
        if r.get("id") == record_id:
            return i, r
    return None, None


def delete_record_by_index(year: int, month: int, index: int, user_id: int = 0) -> dict:
    records = load_records(year, month, user_id)
    deleted = records.pop(index)
    save_records(year, month, records, user_id)
    update_excel(year, month, records, user_id)
    return deleted


def fix_year(year: int, now: datetime) -> tuple[int, bool]:
    if abs(year - now.year) > 1:
        return now.year, True
    return year, False


async def parse_manual_ai(text: str, now: datetime) -> dict | None:
    """自然言語の支出テキストをAIで解析する。書式不要。"""
    profile = get_business_profile()
    profile_line = f"事業プロフィール：{profile}\n" if profile else ""
    prompt = f"""{profile_line}今日の日付：{now.strftime('%Y-%m-%d')}（{now.year}年{now.month}月{now.day}日）

以下のテキストから支出情報を読み取りJSONで返してください。
日付が書かれていない場合は今日の日付を使用してください。
金額が見つからない場合はnullを返してください。

テキスト：{text}

カテゴリは必ず以下から選択：{', '.join(CATEGORIES)}

{{
  "date": "YYYY-MM-DD",
  "name": "品名・店名・内容",
  "amount": 金額（整数。見つからなければnull）,
  "category": "カテゴリ",
  "purpose": "個人 または 仕事"
}}

JSONのみ返してください。"""
    try:
        response = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        m = re.search(r'\{.*\}', response.content[0].text.strip(), re.DOTALL)
        if not m:
            return None
        data = json.loads(m.group())
        if data.get("amount") is None:
            return None
        date_obj = datetime.strptime(data["date"], "%Y-%m-%d")
        return {
            "date":     data["date"],
            "year":     date_obj.year,
            "month":    date_obj.month,
            "name":     data.get("name", "不明"),
            "amount":   int(data["amount"]),
            "category": data.get("category", "その他"),
            "purpose":  data.get("purpose", "個人"),
        }
    except Exception as e:
        logger.error(f"手入力AI解析エラー: {e}")
    return None


def update_excel(year: int, month: int, records: list, user_id: int = 0):
    path = get_excel_path(year, month, user_id)
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "収支一覧"

    hfill  = PatternFill("solid", fgColor="1F4E79")
    ifill  = PatternFill("solid", fgColor="E2EFDA")   # 収入
    efill  = PatternFill("solid", fgColor="FCE4D6")   # 支出・個人
    bizfill = PatternFill("solid", fgColor="D6E4F7")  # 支出・仕事
    hfont  = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col, h in enumerate(["日付", "種別", "用途", "店名・内容", "カテゴリ", "金額（円）", "備考"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = border
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20

    income_total = expense_total = biz_total = personal_total = 0
    for row_idx, r in enumerate(sorted(records, key=lambda x: x.get("date", "")), 2):
        is_income = r.get("type") == "収入"
        is_biz    = r.get("purpose") == "仕事"
        amount    = r.get("amount", 0)
        if is_income:
            income_total += amount
        else:
            expense_total += amount
            if is_biz:
                biz_total += amount
            else:
                personal_total += amount

        if is_income:
            row_fill = ifill
        elif is_biz:
            row_fill = bizfill
        else:
            row_fill = efill

        purpose_label = "" if is_income else ("💼仕事" if is_biz else "🏠個人")
        for col, v in enumerate([
            r.get("date",""), r.get("type",""), purpose_label,
            r.get("name",""), r.get("category",""), amount, r.get("note","")
        ], 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.fill = row_fill
            c.border = border
            if col == 6:
                c.alignment    = Alignment(horizontal="right")
                c.number_format = '#,##0'

    last = len(records) + 2
    for row_offset, label, val, color in [
        (0, "【収入合計】",     income_total,                  "2E75B6"),
        (1, "【支出合計】",     expense_total,                 "C00000"),
        (2, "　うち仕事経費",   biz_total,                     "1F4E79"),
        (3, "　うち個人支出",   personal_total,                "843C0C"),
        (4, "【収支差額】",     income_total - expense_total,  "000000"),
    ]:
        ws.cell(row=last+row_offset, column=4, value=label).font = Font(bold=True)
        c = ws.cell(row=last+row_offset, column=6, value=val)
        c.font = Font(bold=True, color=color)
        c.number_format = '#,##0'

    wb.save(path)


def build_summary(year: int, month: int, records: list, max_items: int = 20) -> str:
    income  = sum(r["amount"] for r in records if r.get("type") == "収入")
    expense = sum(r["amount"] for r in records if r.get("type") == "支出")
    lines   = [
        f"📊 {year}年{month}月の収支\n",
        f"💰 収入合計：¥{income:,}",
        f"💸 支出合計：¥{expense:,}",
        f"📈 収支差額：¥{income - expense:,}\n",
        "--- 明細 ---",
    ]
    for r in sorted(records, key=lambda x: x.get("date", ""))[-max_items:]:
        emoji = "💰" if r["type"] == "収入" else "💸"
        lines.append(f"{emoji} {r['date']} {r['name']} ¥{r['amount']:,} [{r['category']}]")
    if len(records) > max_items:
        lines.append(f"...他{len(records)-max_items}件（Excelで全件確認できます）")
    return "\n".join(lines)


def create_annual_summary(year: int, user_id: int = 0) -> Path:
    """
    国税庁「収支内訳書（一般用）」白色申告フォーマット。
    Sheet1: 収支内訳書（仕事用経費のみ・確定申告提出用）
    Sheet2: 月別売上明細
    Sheet3: 仕事用経費明細
    Sheet4: 個人用支出明細
    """
    output_path = OUTPUT_DIR / f"{year}_確定申告_収支内訳書_{user_id}.xlsx"

    CATEGORY_MAP: dict[str, str] = {
        "交通費":   "旅費交通費",
        "通信費":   "通信費",
        "消耗品費": "消耗品費",
        "光熱費":   "水道光熱費",
        "娯楽費":   "接待交際費",
        "外食費":   "接待交際費",
        "衣服費":   "消耗品費",
        "日用品":   "消耗品費",
        "仕事経費": "雑費",
        "食費":     "消耗品費",
        "医療費":   "雑費",
        "その他":   "雑費",
    }
    EXPENSE_ITEMS: list[str] = [
        "租税公課", "荷造運賃", "水道光熱費", "旅費交通費", "通信費",
        "広告宣伝費", "接待交際費", "損害保険料", "修繕費", "消耗品費",
        "給料賃金", "外注工賃", "減価償却費", "地代家賃", "利子割引料", "雑費",
    ]

    # ── データ集計 ──
    monthly_income:       list[int]       = []
    biz_by_item:          dict[str, int]  = {k: 0 for k in EXPENSE_ITEMS}
    annual_income         = 0
    all_biz_expenses:     list[dict]      = []
    all_personal_expenses: list[dict]     = []

    for month in range(1, 13):
        records = load_records(year, month, user_id)
        inc = sum(r["amount"] for r in records if r.get("type") == "収入")
        monthly_income.append(inc)
        annual_income += inc
        for r in records:
            if r.get("type") != "支出":
                continue
            if r.get("purpose") == "仕事":
                item_key = CATEGORY_MAP.get(r.get("category", "その他"), "雑費")
                biz_by_item[item_key] = biz_by_item.get(item_key, 0) + r.get("amount", 0)
                all_biz_expenses.append(r)
            else:
                all_personal_expenses.append(r)

    total_biz_expense      = sum(biz_by_item.values())
    total_personal_expense = sum(r["amount"] for r in all_personal_expenses)

    # ── スタイル共通 ──
    wb = openpyxl.Workbook()
    thin    = Side(style="thin")
    medium  = Side(style="medium")
    b_thin  = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    hfill   = PatternFill("solid", fgColor="1F4E79")
    sfill   = PatternFill("solid", fgColor="D6E4F7")
    tfill   = PatternFill("solid", fgColor="FFF2CC")
    pfill   = PatternFill("solid", fgColor="E2EFDA")  # 個人用
    center  = Alignment(horizontal="center", vertical="center")
    right_a = Alignment(horizontal="right",  vertical="center")
    left_a  = Alignment(horizontal="left",   vertical="center")
    hfont   = Font(bold=True, color="FFFFFF", size=10)
    bold10  = Font(bold=True, size=10)
    norm10  = Font(size=10)

    def mc(ws, row, col, val, bold=False, color=None):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = '#,##0'; c.alignment = right_a; c.border = b_thin
        c.font = Font(bold=bold, size=10, color=color or "000000")
        return c

    def lc(ws, row, col, val, bold=False, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = left_a; c.border = b_thin
        c.font = Font(bold=bold, size=10)
        if fill: c.fill = fill
        return c

    # ══════════════════════════════════════════════════════
    # Sheet1: 収支内訳書（仕事用・確定申告提出用）
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "収支内訳書（仕事用）"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:E1")
    tc = ws["A1"]
    tc.value = f"{year}年分　収支内訳書（一般用）　白色申告 ／ 仕事用経費"
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"].value = "※ 確定申告書B（第一表）に添付。個人用支出はSheet4参照。"
    ws["A2"].font  = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4

    # 収入
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]; c.value = "■ 1. 収入金額"; c.font = Font(bold=True, size=11)
    c.fill = sfill; c.alignment = left_a; row += 1
    for col, h in enumerate(["科目", "金　額（円）"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = b_thin; c.alignment = center
    row += 1
    lc(ws, row, 1, "売上（収入）金額", bold=True)
    mc(ws, row, 2, annual_income, bold=True, color="2E75B6"); row += 2

    # 売上原価
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]; c.value = "■ 2. 売上原価"; c.font = Font(bold=True, size=11)
    c.fill = sfill; c.alignment = left_a; row += 1
    for col, h in enumerate(["科目", "金　額（円）"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = b_thin; c.alignment = center
    row += 1
    for lbl in ["期首商品（製品）棚卸高", "仕入金額", "小　計", "期末商品（製品）棚卸高", "差引原価"]:
        lc(ws, row, 1, lbl); mc(ws, row, 2, 0); row += 1
    row += 1

    # 経費（仕事用のみ）
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]; c.value = "■ 3. 経費（仕事用のみ）"; c.font = Font(bold=True, size=11)
    c.fill = sfill; c.alignment = left_a; row += 1
    for col, h in enumerate(["経費科目", "金　額（円）", "経費科目", "金　額（円）"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = b_thin; c.alignment = center
    row += 1
    left_items = EXPENSE_ITEMS[:8]; right_items = EXPENSE_ITEMS[8:]
    for i in range(max(len(left_items), len(right_items))):
        if i < len(left_items):
            lc(ws, row, 1, left_items[i]); mc(ws, row, 2, biz_by_item.get(left_items[i], 0))
        if i < len(right_items):
            lc(ws, row, 3, right_items[i]); mc(ws, row, 4, biz_by_item.get(right_items[i], 0))
        row += 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]; c.value = "経費合計（仕事用）"; c.font = bold10; c.fill = tfill
    c.alignment = left_a; c.border = b_thin
    mc(ws, row, 4, total_biz_expense, bold=True, color="C00000"); row += 2

    # 所得金額
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]; c.value = "■ 4. 所得金額"; c.font = Font(bold=True, size=11)
    c.fill = sfill; c.alignment = left_a; row += 1
    for lbl, val, color in [
        ("売上（収入）金額",             annual_income,                       "2E75B6"),
        ("経費合計（仕事用）",            total_biz_expense,                   "C00000"),
        ("青色申告特別控除前の所得金額",  annual_income - total_biz_expense,   "000000"),
    ]:
        lc(ws, row, 1, lbl, bold=True); mc(ws, row, 2, val, bold=True, color=color); row += 1

    row += 1
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].value = f"※ 個人用支出合計：¥{total_personal_expense:,}（確定申告対象外。Sheet4参照）"
    ws[f"A{row}"].font  = Font(italic=True, size=9, color="888888")

    # ══════════════════════════════════════════════════════
    # Sheet2: 月別売上明細
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("月別売上明細")
    for col_letter, w in zip("ABCD", [10, 18, 18, 18]):
        ws2.column_dimensions[col_letter].width = w

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = f"{year}年分　月別収入・支出明細"; ws2.row_dimensions[1].height = 26
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(["月", "収入合計（円）", "仕事経費（円）", "個人支出（円）"], 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = b_thin; c.alignment = center

    ann_biz = ann_per = 0
    for month in range(1, 13):
        records = load_records(year, month, user_id)
        biz = sum(r["amount"] for r in records if r.get("type") == "支出" and r.get("purpose") == "仕事")
        per = sum(r["amount"] for r in records if r.get("type") == "支出" and r.get("purpose") != "仕事")
        ann_biz += biz; ann_per += per
        r_row = month + 3
        lc(ws2, r_row, 1, f"{month}月"); ws2.cell(row=r_row, column=1).alignment = center
        mc(ws2, r_row, 2, monthly_income[month-1])
        mc(ws2, r_row, 3, biz)
        mc(ws2, r_row, 4, per)

    mc(ws2, 16, 2, annual_income,  bold=True, color="2E75B6")
    mc(ws2, 16, 3, ann_biz,        bold=True, color="C00000")
    mc(ws2, 16, 4, ann_per,        bold=True, color="843C0C")
    lc(ws2, 16, 1, "年間合計", bold=True); ws2.cell(row=16, column=1).alignment = center

    # ══════════════════════════════════════════════════════
    # Sheet3: 仕事用経費明細
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("仕事用経費明細")
    for col_letter, w in zip("ABCDEF", [12, 16, 22, 18, 14, 18]):
        ws3.column_dimensions[col_letter].width = w

    ws3.merge_cells("A1:F1")
    ws3["A1"].value = f"{year}年分　仕事用経費明細（確定申告提出用）"
    ws3["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for col, h in enumerate(["日付", "カテゴリ", "内容", "経費科目（申告用）", "金額（円）", "備考"], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.border = b_thin; c.alignment = center

    dr = 4
    for r in sorted(all_biz_expenses, key=lambda x: x.get("date", "")):
        item_key = CATEGORY_MAP.get(r.get("category", "その他"), "雑費")
        for col, val in enumerate([r.get("date",""), r.get("category",""), r.get("name",""),
                                    item_key, r.get("amount",0), r.get("note","")], 1):
            c = ws3.cell(row=dr, column=col, value=val)
            c.border = b_thin; c.font = norm10
            c.alignment = right_a if col == 5 else left_a
            if col == 5: c.number_format = '#,##0'
        dr += 1
    lc(ws3, dr, 3, "仕事経費合計", bold=True)
    c_tot = ws3.cell(row=dr, column=5, value=total_biz_expense)
    c_tot.number_format = '#,##0'; c_tot.font = Font(bold=True, size=10, color="C00000")
    c_tot.border = b_thin; c_tot.alignment = right_a

    # ══════════════════════════════════════════════════════
    # Sheet4: 個人用支出明細
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("個人用支出明細")
    for col_letter, w in zip("ABCDE", [12, 16, 24, 14, 18]):
        ws4.column_dimensions[col_letter].width = w

    ws4.merge_cells("A1:E1")
    ws4["A1"].value = f"{year}年分　個人用支出明細（参考）"
    ws4["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", fgColor="2E7D32")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    for col, h in enumerate(["日付", "カテゴリ", "内容", "金額（円）", "備考"], 1):
        c = ws4.cell(row=3, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="2E7D32"); c.font = hfont
        c.border = b_thin; c.alignment = center

    dr4 = 4
    for r in sorted(all_personal_expenses, key=lambda x: x.get("date", "")):
        for col, val in enumerate([r.get("date",""), r.get("category",""), r.get("name",""),
                                    r.get("amount",0), r.get("note","")], 1):
            c = ws4.cell(row=dr4, column=col, value=val)
            c.border = b_thin; c.font = norm10; c.fill = pfill
            c.alignment = right_a if col == 4 else left_a
            if col == 4: c.number_format = '#,##0'
        dr4 += 1
    ws4.cell(row=dr4, column=3, value="個人支出合計").font = bold10
    ws4.cell(row=dr4, column=3).border = b_thin
    c_per = ws4.cell(row=dr4, column=4, value=total_personal_expense)
    c_per.number_format = '#,##0'; c_per.font = Font(bold=True, size=10, color="2E7D32")
    c_per.border = b_thin; c_per.alignment = right_a

    wb.save(output_path)
    return output_path


# ────────────────────────────────────────────
# AI処理
# ────────────────────────────────────────────

async def analyze_receipt(image_bytes: bytes, media_type: str) -> dict | None:
    """
    レシートを解析し、個人用・仕事用に自動分別して返す。
    混在レシートの場合は records リストに2件返す。
    """
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    business_profile = get_business_profile()
    profile_line = f"【事業プロフィール】\n{business_profile}\n\n" if business_profile else ""
    prompt = f"""{profile_line}このレシート画像を解析してください。
日付が読み取れない場合は"unknown"にしてください。
カテゴリは必ず以下から選んでください：{', '.join(CATEGORIES)}

【重要】金額の読み取りルール：
- 「小計」「合計」「税込合計」「お買い上げ合計」が実際の支払金額です
- 「お預かり」「お預り」は除外してください
- 「お釣り」「おつり」「チェンジ」は除外してください
- ポイント・割引後の実際の支払額を使ってください

【個人/仕事の分別ルール】
- 文具・PC周辺機器・ビジネス書・通信費・交通費（出張）→「仕事」
- 食料品・日用品・医療費・娯楽・衣服（私服）→「個人」
- 同一レシートに混在する場合は品目ごとに金額を分けて2レコード返す
- 判断できない場合は「個人」にする

以下のJSON形式のみで返してください：
{{
  "date": "YYYY-MM-DD",
  "time": "HH:MM（24時間表記。読み取れない場合はnull）",
  "name": "店名または内容",
  "confidence": "high/medium/low",
  "records": [
    {{
      "purpose": "個人 または 仕事",
      "amount": 金額（整数）,
      "category": "カテゴリ",
      "note": "内訳の補足（例：個人用食料品、仕事用文具）"
    }}
  ]
}}

※ 全て個人用なら records に1件（purpose:個人）、全て仕事用なら1件（purpose:仕事）、混在なら2件。
JSONのみ返してください。"""
    try:
        response = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=600,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text",  "text": prompt},
            ]}],
        )
        m = re.search(r'\{.*\}', response.content[0].text.strip(), re.DOTALL)
        if not m:
            return None
        data = json.loads(m.group())
        # 旧形式（amount直返し）との後方互換
        if "amount" in data and "records" not in data:
            data["records"] = [{
                "purpose":  "個人",
                "amount":   data["amount"],
                "category": data.get("category", "その他"),
                "note":     data.get("note", ""),
            }]
        return data
    except Exception as e:
        logger.error(f"AI解析エラー: {e}")
    return None


async def ask_ai_intent(text: str, now: datetime, last_ctx: dict | None) -> dict:
    """メッセージの意図をAIで解析して返す。このチャンネルは家計簿専用なので全メッセージを処理。"""
    ctx_text = "なし（まだ記録していない）"
    if last_ctx:
        r = last_ctx["record"]
        ctx_text = (
            f"日付：{r.get('date')}　店名：{r.get('name')}　"
            f"金額：{r.get('amount')}円　カテゴリ：{r.get('category')}　"
            f"備考：{r.get('note', '')}"
        )

    prev_year  = now.year if now.month > 1 else now.year - 1
    prev_month = now.month - 1 if now.month > 1 else 12

    prompt = f"""あなたは家計簿Botです。ユーザーのメッセージの意図を判定してJSONで返してください。
このチャンネルは家計簿専用なので、全てのメッセージが家計簿に関係すると想定してください。

今日：{now.strftime('%Y-%m-%d')}（{now.year}年{now.month}月）
先月：{prev_year}年{prev_month}月

【直近の記録】
{ctx_text}

【ユーザーのメッセージ】
{text}

【intent一覧と返すJSON】

"edit"   … 直近記録を修正・変更・直す・入れて・書いて・追加して（複数フィールド同時対応）
  → {{"intent":"edit","edits":[{{"field":"amount","value":"800"}},{{"field":"note","value":"接待"}}]}}
  ※ 1フィールドでも必ずeditsリスト形式で返すこと

"delete" … 直近記録を消す・削除・取り消す
  → {{"intent":"delete"}}

"show"   … 収支・明細・一覧を見たい・確認・教えて
  → {{"intent":"show","year":{now.year},"month":{now.month}}}

"category" … 特定カテゴリの合計・いくら
  → {{"intent":"category","category":"食費","year":{now.year},"month":{now.month}}}

"income" … 収入を記録したい
  → {{"intent":"income","amount":15000,"name":"給料"}}

"manual" … 手書き・手入力・写真なしで領収書を入力したい
  → {{"intent":"manual"}}

"summary" … 年間集計・確定申告・まとめ
  → {{"intent":"summary","year":{now.year}}}
- 定期費用の登録（「毎月1日に家賃8万円」「電気代1万円を毎月27日に」）
  → {{"intent":"add_recurring","name":"家賃","amount":80000,"day":1,"category":"その他","purpose":"個人"}}
- 予算設定（「食費の予算3万円」「交通費2万円に設定して」）
  → {{"intent":"set_budget","category":"食費","amount":30000}}
- 過去記録の検索（「セブンの記録」「5000円以上の支出」「食費を全部見せて」）
  → {{"intent":"search","keyword":"セブン","min_amount":0,"max_amount":0,"category":""}}

"help"   … 使い方・ヘルプ
  → {{"intent":"help"}}

"category_advice" … 「これ何費？」「どのカテゴリ？」「経費になる？」などカテゴリの相談
  → {{"intent":"category_advice","description":"相談内容をそのまま"}}

"unknown" … 挨拶など全く関係ないメッセージ
  → {{"intent":"unknown"}}

JSONのみ返してください。"""

    try:
        response = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        m = re.search(r'\{.*\}', response.content[0].text.strip(), re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception as e:
        logger.error(f"AI意図解析エラー: {e}")
    return {"intent": "unknown"}


async def ask_ai_category_advice(description: str, record: dict | None = None) -> str:
    """支出内容からカテゴリのアドバイスをAIが返す。"""
    cats = "、".join(CATEGORIES)
    record_info = ""
    if record:
        record_info = (
            f"\n【直近の記録】\n"
            f"店名：{record.get('name')}　金額：{record.get('amount')}円　"
            f"現在のカテゴリ：{record.get('category')}"
        )

    prompt = f"""あなたは家計簿の専門家です。以下の支出内容に最もふさわしいカテゴリをアドバイスしてください。
{record_info}

【ユーザーの質問・内容】
{description}

【選べるカテゴリ】
{cats}

以下の形式で答えてください：
- おすすめカテゴリ（1〜2個）とその理由を簡潔に
- 確定申告で経費にできるかどうかも一言添える
- 最後に「○○費でよいですか？」と確認する
- 100文字以内でコンパクトに答える"""

    try:
        response = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text.strip()
    except Exception as e:
        logger.error(f"カテゴリアドバイスエラー: {e}")
        return f"カテゴリの候補：{cats}\nどれが近いか教えてください！"


async def ask_ai_reply_edit(text: str, record: dict) -> dict | None:
    """リプライの修正指示をAIで解析する。"""
    record_summary = (
        f"日付：{record.get('date')}　店名：{record.get('name')}　"
        f"金額：{record.get('amount')}円　カテゴリ：{record.get('category')}　"
        f"備考：{record.get('note', '')}"
    )
    prompt = f"""以下の記録に対する修正指示を解析してJSONで返してください。

【現在の記録】
{record_summary}

【修正指示】
{text}

フィールド：amount（金額）/ name（店名）/ date（YYYY-MM-DD）/ category（{', '.join(CATEGORIES)}）/ note（備考・宛名・但し書き）

{{"field": "フィールド名", "value": "新しい値"}}

判断不能の場合は {{"error": "不明"}} を返してください。JSONのみ返してください。"""

    try:
        response = ai_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=128,
            messages=[{"role": "user", "content": prompt}],
        )
        m = re.search(r'\{.*\}', response.content[0].text.strip(), re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception as e:
        logger.error(f"AI修正解析エラー: {e}")
    return None


# ────────────────────────────────────────────
# 修正値のバリデーション・変換（共通処理）
# ────────────────────────────────────────────

def validate_edit(field: str, raw_value: str, now: datetime) -> tuple[any, str, str | None]:
    """
    (new_value, display_value, error_message) を返す。
    error_message が None でなければ失敗。
    """
    field_labels = {
        "amount": "金額", "name": "店名", "date": "日付",
        "category": "カテゴリ", "note": "備考/宛名/但し書き",
    }
    label = field_labels.get(field, field)

    if field == "amount":
        try:
            v = int(str(raw_value).replace(",", "").replace("円", ""))
            return v, f"¥{v:,}", None
        except ValueError:
            return None, None, "❌ 金額の数値が読み取れませんでした。"

    elif field == "date":
        m = re.search(r'(\d{4}-\d{2}-\d{2})|(\d{1,2})[/\-月](\d{1,2})', str(raw_value))
        if not m:
            return None, None, "❌ 日付の形式が読み取れませんでした。例：4/20"
        try:
            if m.group(1):
                v = m.group(1)
            else:
                v = datetime(now.year, int(m.group(2)), int(m.group(3))).strftime("%Y-%m-%d")
            return v, v, None
        except ValueError:
            return None, None, "❌ 日付が正しくありません。"

    elif field == "category":
        cat = next((c for c in CATEGORIES if c in str(raw_value)), None)
        if not cat:
            return None, None, f"❌ カテゴリが見つかりません。\n使えるカテゴリ：{' / '.join(CATEGORIES)}"
        return cat, cat, None

    elif field in ("name", "note"):
        v = str(raw_value).strip()
        return v, v, None

    return None, None, f"❌ フィールド '{field}' は修正できません。"


def apply_edits(user_id: int, edits: list, year: int, month: int, index: int, now: datetime) -> tuple[list, list]:
    """複数フィールドを一括修正。日付変更時はフォルダ・ファイルも移動。
    (成功メッセージリスト, エラーメッセージリスト) を返す。
    """
    records = load_records(year, month, user_id)
    if index < 0 or index >= len(records):
        return [], ["❌ 記録が見つかりませんでした。"]

    field_labels = {"amount":"金額","name":"店名","date":"日付","category":"カテゴリ","note":"備考/宛名/但し書き"}
    ok_lines  = []
    err_lines = []
    changed_fields = set()

    # フィールドを一括更新
    for edit in edits:
        field     = edit.get("field", "")
        raw_value = str(edit.get("value", "")).strip()
        new_value, display_value, err = validate_edit(field, raw_value, now)
        if err:
            err_lines.append(err)
            continue
        old = records[index].get(field, "")
        records[index][field] = new_value
        changed_fields.add(field)
        ok_lines.append(f"{field_labels.get(field, field)}：{old} → {display_value}")
        logger.info(f"✏️ 修正: {field} {old}→{new_value}")

    if not ok_lines:
        return ok_lines, err_lines

    rec          = records[index]
    old_filename = rec.get("receipt_file")
    receipt_dir  = get_month_dir(year, month, user_id) / "receipts"

    # ── 日付が変わった場合：フォルダごと移動 ──
    if "date" in changed_fields:
        new_date_str = rec.get("date", "")
        try:
            new_date_obj = datetime.strptime(new_date_str, "%Y-%m-%d")
        except ValueError:
            err_lines.append("❌ 日付の変換に失敗しました。")
            return ok_lines, err_lines

        new_year  = new_date_obj.year
        new_month = new_date_obj.month

        if new_year != year or new_month != month:
            # 古い月から記録を削除
            records.pop(index)
            save_records(year, month, records, user_id)
            update_excel(year, month, records, user_id)

            # 新しい月に記録を追加
            rec["year"]  = new_year
            rec["month"] = new_month
            new_records  = load_records(new_year, new_month, user_id)
            new_records.append(rec)
            new_idx = len(new_records) - 1
            save_records(new_year, new_month, new_records, user_id)
            update_excel(new_year, new_month, new_records, user_id)

            # レシートファイルを新フォルダに移動＆リネーム
            if old_filename:
                old_path = receipt_dir / old_filename
                if old_path.exists():
                    parts     = old_filename.split("_", 2)
                    date_part = new_date_str.replace("-", "")
                    time_part = parts[1] if len(parts) >= 2 else now.strftime("%H%M%S")
                    ext       = old_filename.rsplit(".", 1)[-1] if "." in old_filename else "jpg"
                    safe_name = re.sub(r'[\/:*?"<>|]', '', rec.get("name", "不明"))[:20]
                    new_filename = f"{date_part}_{time_part}_{safe_name}.{ext}"
                    new_receipt_dir = get_month_dir(new_year, new_month, user_id) / "receipts"
                    new_path = new_receipt_dir / new_filename
                    old_path.rename(new_path)
                    new_records[-1]["receipt_file"] = new_filename
                    save_records(new_year, new_month, new_records, user_id)
                    ok_lines.append(f"📁 {year}/{month:02d}月/receipts/{old_filename}")
                    ok_lines.append(f"　→ {new_year}/{new_month:02d}月/receipts/{new_filename}")
                    logger.info(f"📁 移動: {old_filename} → {new_year}/{new_month:02d}/{new_filename}")

            if user_id in last_record_context:
                last_record_context[user_id] = {
                    "record": new_records[new_idx],
                    "year":   new_year,
                    "month":  new_month,
                    "index":  new_idx,
                }
            return ok_lines, err_lines

        # 同じ年月内での日付変更（ファイル名だけリネーム）
        if old_filename and changed_fields & {"name", "date"}:
            old_path  = receipt_dir / old_filename
            if old_path.exists():
                parts        = old_filename.split("_", 2)
                date_part    = new_date_str.replace("-", "")
                time_part    = parts[1] if len(parts) >= 2 else now.strftime("%H%M%S")
                ext          = old_filename.rsplit(".", 1)[-1] if "." in old_filename else "jpg"
                safe_name    = re.sub(r'[\/:*?"<>|]', '', rec.get("name", "不明"))[:20]
                new_filename = f"{date_part}_{time_part}_{safe_name}.{ext}"
                new_path     = receipt_dir / new_filename
                old_path.rename(new_path)
                records[index]["receipt_file"] = new_filename
                ok_lines.append(f"📁 {old_filename} → {new_filename}")
                logger.info(f"📁 リネーム: {old_filename} → {new_filename}")

    # ── 店名だけ変わった場合：ファイル名リネーム ──
    elif "name" in changed_fields and old_filename:
        old_path = receipt_dir / old_filename
        if old_path.exists():
            parts        = old_filename.split("_", 2)
            date_part    = rec.get("date", "").replace("-", "")
            time_part    = parts[1] if len(parts) >= 2 else now.strftime("%H%M%S")
            ext          = old_filename.rsplit(".", 1)[-1] if "." in old_filename else "jpg"
            safe_name    = re.sub(r'[\/:*?"<>|]', '', rec.get("name", "不明"))[:20]
            new_filename = f"{date_part}_{time_part}_{safe_name}.{ext}"
            new_path     = receipt_dir / new_filename
            old_path.rename(new_path)
            records[index]["receipt_file"] = new_filename
            ok_lines.append(f"📁 {old_filename} → {new_filename}")
            logger.info(f"📁 リネーム: {old_filename} → {new_filename}")

    save_records(year, month, records, user_id)
    update_excel(year, month, records, user_id)

    if user_id in last_record_context:
        last_record_context[user_id]["record"] = records[index]

    return ok_lines, err_lines



def record_confirm_msg(record: dict) -> str:
    purpose_emoji = "💼" if record.get("purpose") == "仕事" else "🏠"
    purpose_label = record.get("purpose", "個人")
    return (
        f"```\n"
        f"日付：{record['date']}\n"
        f"店名：{record['name']}\n"
        f"金額：¥{record['amount']:,}\n"
        f"カテゴリ：{record['category']}\n"
        f"用途：{purpose_emoji} {purpose_label}\n"
        f"備考：{record.get('note', '')}\n"
        f"```\n"
        f"↩️ このメッセージに返信するとその記録を操作できます\n"
        f"||[record_id:{record['id']}]||"
    )


# ────────────────────────────────────────────
# Discordイベント
# ────────────────────────────────────────────

# ────────────────────────────────────────────
# 定期タスク（毎日0時）
# ────────────────────────────────────────────

@tasks.loop(hours=24)
async def daily_check():
    """毎日0時：定期費用の自動記録 + 月初めの予算リセット通知。"""
    if not KAKEIBO_CHANNEL_ID:
        return
    ch = bot.get_channel(KAKEIBO_CHANNEL_ID)
    if not ch:
        return

    now = datetime.now()
    today = now.strftime("%Y-%m-%d")

    for item in get_recurring():
        # 毎月X日に記録
        if item.get("day") != now.day:
            continue
        uid = item.get("user_id", 0)
        record = {
            "id":           make_record_id(today, f"recurring_{item['name']}"),
            "year":         now.year, "month": now.month, "date": today,
            "type":         item.get("type", "支出"),
            "name":         item["name"],
            "amount":       item["amount"],
            "category":     item.get("category", "その他"),
            "purpose":      item.get("purpose", "個人"),
            "note":         "定期費用（自動）",
            "receipt_file": None, "user_id": uid,
        }
        add_record(record)
        await ch.send(
            f"🔄 **定期費用を自動記録しました**\n`{item['name']} / ¥{item['amount']:,} / {item.get('category','その他')}`"
        )
        logger.info(f"🔄 定期記録: {item['name']} ¥{item['amount']:,}")

@daily_check.before_loop
async def before_daily_check():
    await bot.wait_until_ready()
    # 次の0時まで待機
    now   = datetime.now()
    secs  = ((24 - now.hour) * 3600) - (now.minute * 60) - now.second
    await asyncio.sleep(secs % 86400)


@bot.event
async def on_message_delete(message: discord.Message):
    """Botの確認メッセージが削除されたら対応するレコードも削除する。"""
    if message.author != bot.user:
        return
    if KAKEIBO_CHANNEL_ID and message.channel.id != KAKEIBO_CHANNEL_ID:
        return

    content  = message.content or ""
    id_match = re.search(r'\[record_id:([^\]]+)\]', content)
    if not id_match:
        return  # record_id なし or キャッシュなし

    record_id = id_match.group(1)
    date_part = record_id[:8]
    try:
        r_year, r_month = int(date_part[:4]), int(date_part[4:6])
    except ValueError:
        return

    # 全ユーザーのデータから該当レコードを探す
    for user_dir in (KAKEIBO_DIR).iterdir():
        if not user_dir.is_dir():
            continue
        try:
            uid = int(user_dir.name)
        except ValueError:
            continue
        idx, target = find_record_by_id(r_year, r_month, record_id, uid)
        if idx is not None:
            delete_record_by_index(r_year, r_month, idx, uid)
            logger.info(f"🗑️ メッセージ削除により記録を削除: {record_id} user={uid}")
            try:
                await message.channel.send(
                    f"🗑️ メッセージ削除に連動して記録を削除しました。\n"
                    f"`{target.get('date','')} / {target.get('name','')} / ¥{target.get('amount',0):,}`"
                )
            except Exception:
                pass
            return


@bot.event
async def on_ready():
    # 起動時にDriveからデータを復元
    if drive_sync:
        await asyncio.to_thread(drive_sync.restore_all, KAKEIBO_DIR)

    logger.info(f"✅ Bot起動: {bot.user}")
    print(f"✅ 家計簿Bot起動しました: {bot.user}")
    if not KAKEIBO_CHANNEL_ID:
        print("⚠️  警告: KAKEIBO_CHANNEL_ID が未設定です。全チャンネルで動作するためAPI消費が増えます。")
        print("    .env に KAKEIBO_CHANNEL_ID=チャンネルID を設定することを強く推奨します。")
    if not daily_check.is_running():
        daily_check.start()
    if not is_onboarding_done() and KAKEIBO_CHANNEL_ID:
        try:
            ch = bot.get_channel(KAKEIBO_CHANNEL_ID)
            if ch:
                await ch.send(
                    "👋 **家計簿Bot へようこそ！初期設定を行います。**\n\n"
                    "**Q1. 事業の種類を教えてください。**\n"
                    "例：フリーランスエンジニア / 個人事業主（デザイン） / 飲食店 等\n\n"
                    "「スキップ」で省略（後で「設定」と送ると再設定できます）"
                )
        except Exception as e:
            logger.warning(f"初期設定メッセージ送信失敗: {e}")


@bot.event
async def on_message(message: discord.Message):
    if message.author.bot:
        return

    # 航海日誌処理（Notionメモ自動保存）
    if nisshi and await nisshi.handle(message):
        return

    # ── 重複追加の確認応答 ──────────────────────────────
    if user_id in pending_duplicate and text_raw == "重複追加":
        pd = pending_duplicate.pop(user_id)
        # 保存処理を続行（重複チェックをスキップして記録）
        date_obj     = pd["date_obj"]
        result       = pd["result"]
        rec_list     = pd["rec_list"]
        store_name   = pd["store_name"]
        fname        = pd["fname"]
        image_bytes  = pd["image_bytes"]
        media_type   = pd["media_type"]
        user_comment = pd["user_comment"]
        attachment   = pd["attachment"]
        saved_records = []
        receipt_saved = False
        for i, rec_item in enumerate(rec_list):
            purpose = rec_item.get("purpose", "個人")
            amount  = rec_item.get("amount", 0)
            cat     = rec_item.get("category", "その他")
            note    = user_comment if (user_comment and i == 0) else rec_item.get("note", "")
            record  = {
                "id": make_record_id(result["date"], f"{attachment.filename}_{i}_dup"),
                "year": date_obj.year, "month": date_obj.month,
                "date": result["date"], "type": "支出",
                "name": store_name, "amount": amount,
                "category": cat, "purpose": purpose,
                "note": note, "receipt_file": fname if not receipt_saved else None,
            }
            receipt_saved = True
            add_record(record)
            saved_records.append(record)
        await message.channel.send(f"✅ 重複を承認して記録しました！\n{record_confirm_msg(saved_records[0])}")
        return
    cfg = load_config()
    if "_pending_alert" in cfg:
        alert_msg = cfg.pop("_pending_alert")
        save_config(cfg)
        try:
            await message.channel.send(alert_msg)
        except Exception:
            pass

    # ── チャンネル制限 ──
    if KAKEIBO_CHANNEL_ID and message.channel.id != KAKEIBO_CHANNEL_ID:
        await bot.process_commands(message)
        return

    # ── 経理モードの切り替えを最優先で処理 ──
    text_raw = message.content.strip()
    user_id  = message.author.id  # ← K1修正: user_id を先に定義
    if text_raw in ["経理モード", "けいりもーど", "経理", "会計モード", "終了", "おわり", "exit"]:
        if text_raw in ["終了", "おわり", "exit"] or accounting_mode.get(user_id):
            accounting_mode.pop(user_id, None)
            await message.channel.send(
                "📒 経理モードを終了しました。通常の家計簿モードに戻ります。"
            )
        else:
            accounting_mode[user_id] = True
            await message.channel.send(
                "📊 **経理モードに入りました。**\n\n"
                "税務・経理に関する質問をどうぞ。\n"
                "例：「接待交際費の上限は？」「インボイス登録すべき？」「家事按分の計算方法」\n\n"
                "💡 簡単な質問はHaiku、専門的な質問はSonnetが自動で回答します。\n"
                "「終了」で家計簿モードに戻ります。"
            )
        return

    # ── 経理モード中は質問をAIアドバイザーに転送 ──
    if accounting_mode.get(user_id):
        if not text_raw:
            return
        await message.channel.send("🤔 調べています...")
        answer, model_used = await ask_accounting(text_raw, user_id)
        await message.channel.send(
            f"📊 **経理AIの回答** （{model_used}）\n\n{answer}\n\n"
            f"*⚠️ 重要な判断は税理士にご確認ください。*"
        )
        return

    text    = message.content.strip()
    now     = datetime.now()

    # ══════════════════════════════════════
    # ① オンボーディング（初回ヒアリング）
    # ══════════════════════════════════════
    if user_id in onboarding_state:
        state = onboarding_state[user_id]
        step  = state.get("step", 1)

        if text in ["キャンセル", "スキップ"]:
            cfg = load_config()
            cfg["onboarding_done"] = True
            save_config(cfg)
            del onboarding_state[user_id]
            await message.channel.send("⏭️ 初期設定をスキップしました。「設定」と送ると再度ヒアリングできます。")
            return

        if step == 1:
            state["business_type"] = text
            state["step"] = 2
            onboarding_state[user_id] = state
            await message.channel.send(
                "**Q2. 主な仕事経費を教えてください。**\n例：PC機材・交通費・外食（接待）・書籍など\n「スキップ」で省略"
            )
            return

        if step == 2:
            state["main_expenses"] = text
            state["step"] = 3
            onboarding_state[user_id] = state
            await message.channel.send(
                "**Q3. 自宅兼仕事場ですか？**\n「はい」→ 光熱費・家賃を按分対象にします\n「いいえ」→ 按分なし"
            )
            return

        if step == 3:
            home_office = text in ["はい", "yes", "YES", "うん", "おk", "ok", "OK"]
            profile = (
                f"事業内容：{state.get('business_type', '')}。"
                f"主な仕事経費：{state.get('main_expenses', '')}。"
                f"自宅兼仕事場：{'あり（光熱費・家賃は按分対象）' if home_office else 'なし'}。"
            )
            cfg = load_config()
            cfg["business_profile"] = profile
            cfg["home_office"]       = home_office
            cfg["onboarding_done"]   = True
            save_config(cfg)
            del onboarding_state[user_id]
            await message.channel.send(
                f"✅ 設定完了！\n\n```\n{profile}\n```\n\n"
                f"これでレシートの個人/仕事分別の精度が上がります。\n変更したい場合は「設定」と送ってください。"
            )
            return

    # ══════════════════════════════════════
    # ② 削除確認待ち
    # ══════════════════════════════════════
    if user_id in pending_delete:
        state = pending_delete.pop(user_id)
        if text in ["はい", "yes", "YES", "うん", "おk", "ok", "OK"]:
            deleted = delete_record_by_index(state["year"], state["month"], state["index"], user_id)
            receipt_file = deleted.get("receipt_file")
            if receipt_file:
                (get_month_dir(state["year"], state["month"], user_id) / "receipts" / receipt_file).unlink(missing_ok=True)
            logger.info(f"🗑️ 削除: {deleted.get('name')} ¥{deleted.get('amount',0):,}")
            await message.channel.send(
                f"🗑️ 削除しました！\n```\n日付：{deleted.get('date')}\n店名：{deleted.get('name')}\n金額：¥{deleted.get('amount',0):,}\n```"
            )
        else:
            await message.channel.send("❌ キャンセルしました。")
        return

    # ── リプライは全ての保留状態をクリアして優先処理 ──
    if message.reference:
        for _d in [pending_confirm, pending_manual, pending_date]:
            _d.pop(user_id, None)

    # ══════════════════════════════════════
    # ③ 手入力確認待ち
    # ══════════════════════════════════════
    if user_id in pending_confirm:
        # 5分以上前の保留はタイムアウトとして破棄
        created = pending_confirm[user_id].get("created_at", now.timestamp())
        if now.timestamp() - created > 300:
            pending_confirm.pop(user_id, None)
            await message.channel.send("⏱️ 確認待ちがタイムアウトしました。最初からやり直してください。")
        
    if user_id in pending_confirm:
        state = pending_confirm.pop(user_id)
        if text in ["はい", "yes", "YES", "うん", "おk", "ok", "OK"]:
            record        = state["record"]
            receipt_bytes = state.get("receipt_bytes")
            sent_at       = state.get("sent_at", now)
            if receipt_bytes:
                fname = make_receipt_filename(record["date"], record["name"],
                                              state.get("receipt_ext","receipt.jpg"), None, sent_at)
                record["receipt_file"] = fname
                (get_month_dir(record["year"], record["month"], user_id) / "receipts" / fname).write_bytes(receipt_bytes)
            add_record(record)
            set_last_context(user_id, record)
            logger.info(f"✅ 手入力記録: {record['name']} ¥{record['amount']:,}")
        # 混在レシートの2件目以降も保存
        extra = state.get("extra_records", [])
        store_name = state.get("store_name", record.get("name","不明"))
        for j, ex in enumerate(extra):
            ex_rec = {
                "id": make_record_id(record["date"], f"extra{j}"),
                "year": record["year"], "month": record["month"],
                "date": record["date"], "type": "支出",
                "name": store_name, "amount": ex.get("amount",0),
                "category": ex.get("category","その他"),
                "purpose": ex.get("purpose","個人"),
                "note": ex.get("note",""), "receipt_file": None, "user_id": user_id,
            }
            add_record(ex_rec)
        logger.info(f"✅ extra_records {len(extra)}件 保存")
        await message.channel.send(f"✅ 記録しました！\n{record_confirm_msg(record)}")
    # ══════════════════════════════════════
    # ④ 手入力受付中
    # ══════════════════════════════════════
    if user_id in pending_manual:
        created = pending_manual[user_id].get("sent_at", now)
        if isinstance(created, datetime) and (now - created).total_seconds() > 300:
            pending_manual.pop(user_id, None)
            await message.channel.send("⏱️ 手入力がタイムアウトしました。もう一度「手入力」から始めてください。")
            return
        state = pending_manual.pop(user_id)
        if "キャンセル" in text or "やめ" in text:
            await message.channel.send("❌ キャンセルしました。")
            return
        await message.channel.send("🔍 読み取り中...")
        parsed = await parse_manual_ai(text, now)
        if not parsed:
            await message.channel.send(
                "❓ 金額が読み取れませんでした。\n"
                "例：「ジュース160円」「昨日スーパーで3200円」\n「キャンセル」で中止"
            )
            pending_manual[user_id] = state
            return
        purpose_emoji = "💼" if parsed["purpose"] == "仕事" else "🏠"
        record = {
            "id": make_record_id(parsed["date"], "manual"),
            "year": parsed["year"], "month": parsed["month"],
            "date": parsed["date"], "type": "支出",
            "name": parsed["name"], "amount": parsed["amount"],
            "category": parsed["category"], "purpose": parsed["purpose"],
            "note": "手入力", "receipt_file": None,
            "user_id": user_id,
        }
        add_record(record)
        set_last_context(user_id, record)
        logger.info(f"✅ 手入力記録: {record['name']} ¥{record['amount']:,}")
        await message.channel.send(f"✅ 記録しました！\n{record_confirm_msg(record)}")
        return
    # ══════════════════════════════════════
    # ⑤ 日付確認待ち
    # ══════════════════════════════════════
    if user_id in pending_date:
        sent = pending_date[user_id].get("sent_at", now)
        if isinstance(sent, datetime) and (now - sent).total_seconds() > 300:
            pending_date.pop(user_id, None)
            await message.channel.send("⏱️ 日付入力がタイムアウトしました。最初からやり直してください。")
            return
        state = pending_date.pop(user_id)
        if "キャンセル" in text or "やめ" in text:
            await message.channel.send("❌ 記録をキャンセルしました。")
            return
        date_match = re.search(r'(?:(\d{1,2})月)?(\d{1,2})日', text)
        if date_match:
            try:
                parsed_date = datetime(now.year,
                                       int(date_match.group(1)) if date_match.group(1) else now.month,
                                       int(date_match.group(2)))
                state["record"]["date"]  = parsed_date.strftime("%Y-%m-%d")
                state["record"]["year"]  = parsed_date.year
                state["record"]["month"] = parsed_date.month
            except ValueError:
                await message.channel.send("❌ 日付が正しくありません。キャンセルしました。")
                return
        else:
            await message.channel.send("❓ 「4月20日」のように入力してください。「キャンセル」で中止")
            pending_date[user_id] = state
            return

        record        = state["record"]
        receipt_bytes = state.get("receipt_bytes")
        if receipt_bytes:
            fname = make_receipt_filename(record["date"], record["name"],
                                          state.get("receipt_ext","receipt.jpg"),
                                          state.get("receipt_time"), state.get("sent_at", now))
            record["receipt_file"] = fname
            (get_month_dir(record["year"], record["month"], user_id) / "receipts" / fname).write_bytes(receipt_bytes)

        add_record(record)
        set_last_context(user_id, record)
        logger.info(f"✅ 記録（日付確定後）: {record['name']} ¥{record['amount']:,}")
        # 混在レシートの2件目以降も保存
        extra = state.get("extra_records", [])
        store_name = state.get("store_name", record.get("name","不明"))
        for j, ex in enumerate(extra):
            ex_rec = {
                "id": make_record_id(record["date"], f"extra{j}"),
                "year": record["year"], "month": record["month"],
                "date": record["date"], "type": "支出",
                "name": store_name, "amount": ex.get("amount",0),
                "category": ex.get("category","その他"),
                "purpose": ex.get("purpose","個人"),
                "note": ex.get("note",""), "receipt_file": None, "user_id": user_id,
            }
            add_record(ex_rec)
        logger.info(f"✅ extra_records {len(extra)}件 保存")
        await message.channel.send(f"✅ 記録しました！\n{record_confirm_msg(record)}")
        return

    # ══════════════════════════════════════
    # ⑥ リプライ処理（特定記録への操作）
    # ══════════════════════════════════════
    if message.reference:
        try:
            ref_msg = await message.channel.fetch_message(message.reference.message_id)
        except Exception:
            await message.channel.send("❌ 返信先のメッセージが見つかりませんでした。")
            return
        else:
            id_match = re.search(r'\[record_id:([^\]]+)\]', ref_msg.content)
            if id_match:
                record_id = id_match.group(1)
                date_part = record_id[:8]
                try:
                    r_year, r_month = int(date_part[:4]), int(date_part[4:6])
                except ValueError:
                    r_year, r_month = now.year, now.month

                idx, target = find_record_by_id(r_year, r_month, record_id, user_id)
                if idx is None:
                    await message.channel.send("❌ 既に削除済みか、記録が見つかりませんでした。")
                    return

                # リプライ内容をAIで解析
                intent = await ask_ai_intent(text, now, {"record": target, "year": r_year, "month": r_month, "index": idx})

                if intent.get("intent") == "delete":
                    pending_delete[user_id] = {"year": r_year, "month": r_month, "index": idx, "record": target}
                    await message.channel.send(
                        f"⚠️ 本当に削除しますか？\n```\n日付：{target.get('date')}\n店名：{target.get('name')}\n"
                        f"金額：¥{target.get('amount',0):,}\nカテゴリ：{target.get('category')}\n```\n"
                        f"「はい」で削除 / 「いいえ」でキャンセル"
                    )
                    return

                if intent.get("intent") == "edit":
                    edits = intent.get("edits", [])
                    if not edits and intent.get("field"):
                        edits = [{"field": intent.get("field"), "value": intent.get("value","")}]
                    ok_lines, err_lines = apply_edits(user_id, edits, r_year, r_month, idx, now)
                    if ok_lines:
                        await message.channel.send(f"✅ 修正しました！\n```\n" + "\n".join(ok_lines) + "\n```")
                    for err in err_lines:
                        await message.channel.send(err)
                    return

                # その他のリプライ操作案内
                await message.channel.send(
                    "💬 この記録に対してできる操作：\n"
                    "　🗑️ 削除したい → 「消して」「取り消し」\n"
                    "　✏️ 修正したい → 「金額を800円に」「備考に交通費と入れて」など"
                )
                return

    # ══════════════════════════════════════
    # ⑦ レシート画像
    # ══════════════════════════════════════
    if message.attachments:
        for attachment in message.attachments:
            if not any(attachment.filename.lower().endswith(ext)
                       for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"]):
                continue

            # レシートを送ったら全保留状態をクリア
            cleared = []
            for d, name in [(pending_confirm, "確認"), (pending_manual, "手入力"), (pending_date, "日付入力")]:
                if user_id in d:
                    d.pop(user_id, None)
                    cleared.append(name)
            if cleared:
                await message.channel.send(f"⚠️ 保留中の{'/'.join(cleared)}をキャンセルしてレシートを処理します。")
            await message.channel.send("📸 レシートを読み取り中...")
            logger.info(f"📸 レシート受信: {attachment.filename} user={user_id}")

            async with aiohttp.ClientSession() as session:
                async with session.get(attachment.url) as resp:
                    image_bytes = await resp.read()
            if len(image_bytes) > 5 * 1024 * 1024:
                await message.channel.send("❌ 画像が大きすぎます（上限5MB）。縮小して送ってください。")
                return

            ext = attachment.filename.lower().rsplit(".", 1)[-1]
            media_type = {"jpg":"image/jpeg","jpeg":"image/jpeg","png":"image/png",
                          "webp":"image/webp","gif":"image/gif"}.get(ext, "image/jpeg")

            result = await analyze_receipt(image_bytes, media_type)
            if not result:
                logger.warning(f"❌ AI解析失敗: {attachment.filename}")
                pending_manual[user_id] = {"receipt_bytes": image_bytes,
                                           "receipt_ext": attachment.filename, "sent_at": now}
                await message.channel.send(
                    "❌ 読み取りに失敗しました。手入力で登録できます。\n\n"
                    "書式：`月/日 店名 金額 カテゴリ`\n例：`4/20 薬局 1200 医療費`\n「キャンセル」で中止"
                )
                return

            # 日付不明
            if result.get("date") in ("unknown", "", None):
                user_comment = text.strip()
                rec_list = result.get("records", [{"purpose":"個人","amount":0,"category":"その他","note":""}])
                first_rec = rec_list[0]
                record = {
                    "id": make_record_id("00000000", attachment.filename),
                    "year": now.year, "month": now.month, "date": "",
                    "type": "支出", "name": result.get("name","不明"),
                    "amount": first_rec.get("amount",0), "category": first_rec.get("category","その他"),
                    "purpose": first_rec.get("purpose","個人"),
                    "note": user_comment if user_comment else first_rec.get("note",""), "receipt_file": None,
                }
                pending_date[user_id] = {
                    "record": record, "receipt_bytes": image_bytes,
                    "receipt_time": result.get("time"), "receipt_ext": attachment.filename, "sent_at": now,
                    "extra_records": rec_list[1:],  # 混在レシートの残り分
                    "store_name": result.get("name","不明"),
                }
                purpose_emoji = "💼" if record["purpose"] == "仕事" else "🏠"
                await message.channel.send(
                    f"📅 日付が読み取れませんでした。何月何日のレシートですか？（例：4月20日）\n「キャンセル」で中止\n\n"
                    f"```\n店名：{record['name']}\n金額：¥{record['amount']:,}\nカテゴリ：{record['category']}\n用途：{purpose_emoji} {record['purpose']}\n```"
                )
                return

            date_obj = datetime.strptime(result["date"], "%Y-%m-%d")
            original_year = date_obj.year
            fixed_year, was_fixed = fix_year(original_year, now)
            if was_fixed:
                logger.warning(f"⚠️ 年補正: {original_year}→{fixed_year} ファイル={attachment.filename}")
                date_obj       = date_obj.replace(year=fixed_year)
                result["date"] = date_obj.strftime("%Y-%m-%d")
                await message.channel.send(
                    f"⚠️ 年が `{original_year}年` と読み取られましたが **{fixed_year}年** に自動補正しました。\n"
                    f"違う場合はこのメッセージに返信して修正してください。"
                )

            if result.get("confidence") == "low":
                await message.channel.send("⚠️ 読み取り精度が低いです。内容を確認してください。")

            fname = make_receipt_filename(result["date"], result.get("name","不明"),
                                          attachment.filename, result.get("time"), now)
            user_comment = text.strip()
            store_name   = result.get("name","不明")

            # records リストから各レコードを保存（個人/仕事の分別）
            rec_list      = result.get("records", [{"purpose":"個人","amount":0,"category":"その他","note":""}])
            saved_records = []
            receipt_saved = False

            # ── 重複チェック ──────────────────────────────
            existing = load_records(date_obj.year, date_obj.month, user_id)
            total_new = sum(r.get("amount", 0) for r in rec_list)
            duplicates = [
                r for r in existing
                if r.get("date") == result["date"]
                and r.get("name") == store_name
                and abs(r.get("amount", 0) - total_new) < 10  # 端数誤差10円以内
            ]
            if duplicates:
                await message.channel.send(
                    f"⚠️ **重複の可能性があります**\n"
                    f"`{result['date']} {store_name} ¥{total_new:,}` はすでに記録済みです。\n"
                    f"追加する場合は「重複追加」と送信してください。"
                )
                pending_duplicate[user_id] = {
                    "result": result, "date_obj": date_obj, "rec_list": rec_list,
                    "store_name": store_name, "fname": fname, "image_bytes": image_bytes,
                    "media_type": media_type, "user_comment": user_comment,
                    "attachment": attachment
                }
                return
            # ─────────────────────────────────────────────

            for i, rec_item in enumerate(rec_list):
                purpose = rec_item.get("purpose", "個人")
                amount  = rec_item.get("amount", 0)
                cat     = rec_item.get("category", "その他")
                note    = user_comment if (user_comment and i == 0) else rec_item.get("note", "")

                record = {
                    "id":           make_record_id(result["date"], f"{attachment.filename}_{i}"),
                    "year":         date_obj.year,
                    "month":        date_obj.month,
                    "date":         result["date"],
                    "type":         "支出",
                    "name":         store_name,
                    "amount":       amount,
                    "category":     cat,
                    "purpose":      purpose,
                    "note":         note,
                    "receipt_file": fname if not receipt_saved else None,
                }
                receipt_saved = True
                add_record(record)
                saved_records.append(record)
                logger.info(f"✅ 記録: {store_name} ¥{amount:,} [{cat}][{purpose}]")

            set_last_context(user_id, saved_records[-1])
            # K2修正: user_idを渡す / K3修正: Railwayではレシートをローカル保存しない
            receipt_path = get_month_dir(date_obj.year, date_obj.month, user_id) / "receipts" / fname
            IS_RAILWAY = os.environ.get("RAILWAY_ENVIRONMENT") is not None
            if IS_RAILWAY:
                try:
                    import nisshi as _nisshi
                    import asyncio
                    drive_url = await asyncio.to_thread(
                        _nisshi.upload_to_drive, image_bytes, fname, media_type
                    )
                    logger.info(f"📸 レシートをDriveに保存: {drive_url}")
                except Exception as e:
                    logger.warning(f"Drive保存失敗（ローカルに保存）: {e}")
                    receipt_path.write_bytes(image_bytes)
            else:
                receipt_path.write_bytes(image_bytes)

            # 混在レシートの場合は内訳を表示
            if len(saved_records) == 1:
                await message.channel.send(f"✅ 記録しました！\n{record_confirm_msg(saved_records[0])}")
            else:
                lines = [f"✅ {store_name} のレシートを個人/仕事に自動分別して記録しました！\n"]
                for r in saved_records:
                    emoji = "💼" if r["purpose"] == "仕事" else "🏠"
                    lines.append(f"{emoji} {r['purpose']}用：¥{r['amount']:,}（{r['category']}）")
                total = sum(r["amount"] for r in saved_records)
                lines.append(f"\n合計：¥{total:,}")
                await message.channel.send("\n".join(lines))

            # カテゴリが「その他」の場合は自動でアドバイス
            for r in saved_records:
                if r["category"] == "その他":
                    advice = await ask_ai_category_advice(
                        f"{r['name']} ¥{r['amount']}円", r
                    )
                    await message.channel.send(f"💬 カテゴリが「その他」になりました。\n\n{advice}")
        return

    # ══════════════════════════════════════
    # ⑧ 全テキストメッセージ → AIで意図判定
    # ══════════════════════════════════════
    # ── キーワード直接判定（AIコスト節約・確実動作）──
    text_stripped = text.strip()
    if text_stripped in ["経理相談", "税務相談", "経費相談"]:
        accounting_mode[user_id] = True
        await message.channel.send(
            "📊 **経理モードに入りました。**\n\n"
            "税務・経理に関する質問をどうぞ。\n"
            "「終了」で家計簿モードに戻ります。"
        )
        return

    if text_stripped in ["設定", "せってい", "初期設定", "再設定"]:
        onboarding_state[user_id] = {"step": 1}
        await message.channel.send(
            "⚙️ **事業プロフィールを設定します。**\n\n"
            "**Q1. 事業の種類を教えてください。**\n"
            "例：フリーランスエンジニア / 個人事業主（デザイン） / 飲食店 等\n\n"
            "「スキップ」で省略"
        )
        return

    if text_stripped in ["手入力", "てにゅうりょく", "直接入力", "入力"]:
        pending_manual[user_id] = {"receipt_bytes": None, "receipt_ext": None, "sent_at": now}
        await message.channel.send(
            "✏️ 何を買いましたか？普通に教えてください。\n\n"
            "例：「ジュース160円」\n"
            "　　「昨日スーパーで3200円」\n"
            "　　「5/15 ランチ代 1800円 仕事」\n\n"
            "「キャンセル」で中止"
        )
        return

    # ── 初回ユーザー検知：設定未完了なら先にオンボーディング ──
    # ただし金額を含む支出っぽいメッセージは直接手入力として処理する
    import re as _re
    has_amount = bool(_re.search(r'[0-9０-９]+', text))
    if not is_onboarding_done() and user_id not in onboarding_state and not has_amount:
        onboarding_state[user_id] = {"step": 1}
        await message.channel.send(
            "👋 はじめまして！まず簡単な初期設定をさせてください（3問だけ）。\n\n"
            "**Q1. 事業の種類を教えてください。**\n"
            "例：フリーランスエンジニア / 個人事業主（デザイン） / 飲食店 等\n\n"
            "「スキップ」で省略"
        )
        return

    # ── 金額を含む短いメッセージ → 直接手入力として処理 ──
    import re as _re2
    INCOME_WORDS = ["もらった", "入った", "収入", "給料", "売上", "振込", "報酬", "入金", "稼いだ", "もらえた"]
    QUERY_WORDS  = ["今月", "先月", "いくら", "教えて", "見せて", "まとめ", "収支", "明細", "時間", "時", "分",
                    "月の", "番号", "電話", "ID", "パス", "コード",
                    "持ってきて", "持って来て", "送って", "貸して", "払って",
                    "明日", "あした", "来週", "予定", "かな", "かも", "だろう"]
    is_short_expense = (
        len(text) < 40
        and bool(_re2.search(r'[0-9０-９]{2,}', text))
        and not any(kw in text for kw in QUERY_WORDS)
        and not any(kw in text for kw in INCOME_WORDS)
    )
    if is_short_expense and user_id not in pending_manual:
        await message.channel.send("🔍 読み取り中...")
        parsed = await parse_manual_ai(text, now)
        if parsed:
            purpose_emoji = "💼" if parsed["purpose"] == "仕事" else "🏠"
            record = {
                "id": make_record_id(parsed["date"], "manual"),
                "year": parsed["year"], "month": parsed["month"],
                "date": parsed["date"], "type": "支出",
                "name": parsed["name"], "amount": parsed["amount"],
                "category": parsed["category"], "purpose": parsed["purpose"],
                "note": "手入力", "receipt_file": None, "user_id": user_id,
            }
            add_record(record)
            set_last_context(user_id, record)
            logger.info(f"✅ 直接入力記録: {record['name']} ¥{record['amount']:,}")
            await message.channel.send(f"✅ 記録しました！\n{record_confirm_msg(record)}")
            return

    last_ctx = last_record_context.get(user_id)
    intent   = await ask_ai_intent(text, now, last_ctx)
    action   = intent.get("intent", "unknown")
    logger.info(f"🤖 intent={action} user={user_id} text={text[:40]}")

    # 直近記録の修正（複数フィールド対応・即反映）
    if action == "edit":
        if not last_ctx:
            await message.channel.send("📭 まだ記録がありません。レシートを送るか「手入力」で記録してください。")
            return
        edits = intent.get("edits", [])
        # 旧形式（field/value）との後方互換
        if not edits and intent.get("field"):
            edits = [{"field": intent.get("field"), "value": intent.get("value","")}]
        if not edits:
            await message.channel.send("❓ 修正内容が読み取れませんでした。")
            return
        ok_lines, err_lines = apply_edits(
            user_id, edits,
            last_ctx["year"], last_ctx["month"], last_ctx["index"], now
        )
        if ok_lines:
            await message.channel.send(f"✅ 修正しました！\n```\n" + "\n".join(ok_lines) + "\n```")
        for err in err_lines:
            await message.channel.send(err)
        return

    # 直近記録の削除
    if action == "delete":
        if not last_ctx:
            await message.channel.send("📭 まだ記録がありません。")
            return
        target = last_ctx["record"]
        pending_delete[user_id] = {
            "year": last_ctx["year"], "month": last_ctx["month"],
            "index": last_ctx["index"], "record": target,
        }
        await message.channel.send(
            f"⚠️ この記録を削除しますか？\n```\n日付：{target.get('date')}\n店名：{target.get('name')}\n"
            f"金額：¥{target.get('amount',0):,}\nカテゴリ：{target.get('category')}\n```\n"
            f"「はい」で削除 / 「いいえ」でキャンセル"
        )
        return

    # 収支表示
    if action == "show":
        t_year, t_month = intent.get("year", now.year), intent.get("month", now.month)
        records = load_records(t_year, t_month, user_id)
        if not records:
            await message.channel.send(f"📭 {t_year}年{t_month}月の記録はまだありません。")
            return
        await message.channel.send(build_summary(t_year, t_month, records))
        return

    # カテゴリ別集計
    if action == "category":
        cat     = intent.get("category", "")
        t_year  = intent.get("year",  now.year)
        t_month = intent.get("month", now.month)
        records = load_records(t_year, t_month, user_id)
        items   = [r for r in records if r.get("category") == cat and r.get("type") == "支出"]
        total   = sum(r["amount"] for r in items)
        lines   = [f"📂 {t_year}年{t_month}月の【{cat}】\n合計：¥{total:,}\n"]
        for r in sorted(items, key=lambda x: x.get("date","")):
            lines.append(f"💸 {r['date']} {r['name']} ¥{r['amount']:,}")
        await message.channel.send("\n".join(lines))
        return

    # 収入記録
    if action == "income":
        amount = int(str(intent.get("amount", 0)).replace(",",""))
        name   = str(intent.get("name", "収入")).strip() or "収入"
        record = {
            "id": make_record_id(now.strftime("%Y-%m-%d"), "income"),
            "year": now.year, "month": now.month, "date": now.strftime("%Y-%m-%d"),
            "type": "収入", "name": name, "amount": amount, "category": "収入", "note": "",
            "user_id": user_id,
        }
        add_record(record)
        set_last_context(user_id, record)
        logger.info(f"💰 収入: {name} ¥{amount:,}")
        await message.channel.send(f"✅ 収入を記録しました！\n{record_confirm_msg(record)}")
        return

    # 手入力モード
    if action == "manual":
        pending_manual[user_id] = {"receipt_bytes": None, "receipt_ext": None, "sent_at": now}
        await message.channel.send(
            "✏️ 何を買いましたか？普通に教えてください。\n\n"
            "例：「ジュース160円」\n"
            "　　「昨日スーパーで3200円」\n"
            "　　「5/15 ランチ代 1800円 仕事」\n\n"
            "「キャンセル」で中止"
        )
        return

    # 定期費用の登録
    if action == "add_recurring":
        item = {
            "name":     intent.get("name", ""),
            "amount":   intent.get("amount", 0),
            "day":      intent.get("day", 1),
            "category": intent.get("category", "その他"),
            "purpose":  intent.get("purpose", "個人"),
            "type":     "支出",
            "user_id":  user_id,
        }
        if not item["name"] or not item["amount"]:
            await message.channel.send("❓ 名前と金額が読み取れませんでした。\n例：「毎月1日に家賃8万円」")
            return
        items = get_recurring()
        items = [i for i in items if not (i["name"] == item["name"] and i["user_id"] == user_id)]
        items.append(item)
        set_recurring(items)
        await message.channel.send(
            f"🔄 定期費用を登録しました！\n"
            f"`毎月{item['day']}日 / {item['name']} / ¥{item['amount']:,} / {item['category']}`\n"
            f"毎月自動で記録されます。"
        )
        return

    # 予算設定
    if action == "set_budget":
        cat    = intent.get("category", "")
        amount = intent.get("amount", 0)
        if not cat or not amount:
            await message.channel.send("❓ カテゴリと金額が読み取れませんでした。\n例：「食費の予算3万円」")
            return
        budgets = get_budgets()
        budgets[cat] = amount
        set_budgets(budgets)
        await message.channel.send(f"💰 **{cat}**の月予算を **¥{amount:,}** に設定しました！\n80%超えたら通知します。")
        return

    # 過去記録の検索
    if action == "search":
        keyword    = intent.get("keyword", "")
        min_amount = intent.get("min_amount", 0)
        max_amount = intent.get("max_amount", 0)
        category   = intent.get("category", "")
        results    = search_records(user_id, keyword, min_amount, max_amount, category)
        if not results:
            await message.channel.send("🔍 該当する記録が見つかりませんでした。")
            return
        lines = [f"🔍 **{len(results)}件**見つかりました\n"]
        for r in results[:10]:
            purpose_e = "💼" if r.get("purpose") == "仕事" else "🏠"
            lines.append(f"`{r.get('date','')}` {r.get('name','')} **¥{r.get('amount',0):,}** {r.get('category','')} {purpose_e}")
        if len(results) > 10:
            lines.append(f"…他 {len(results)-10} 件（Excelで確認してください）")
        await message.channel.send("\n".join(lines))
        return

    # 年間集計
    if action == "summary":
        s_year = intent.get("year", now.year)
        await message.channel.send(f"📊 {s_year}年の確定申告用ファイルを作成中...")
        path = create_annual_summary(s_year, user_id)
        await message.channel.send(
            f"✅ 作成しました！\n`{path}`\n\n白色申告の収支内訳書フォーマットで保存しました。",
            file=discord.File(str(path)),
        )
        return

    # カテゴリ相談
    if action == "category_advice":
        description = intent.get("description", text)
        advice = await ask_ai_category_advice(description, last_ctx["record"] if last_ctx else None)
        await message.channel.send(f"💬 {advice}")
        return

    # ヘルプ
    if action == "help":
        await message.channel.send(
            "📖 **家計簿Botの使い方**\n\n"
            "📸 **レシート記録** → 画像を送るだけ\n"
            "💰 **収入記録** → 「給料15万入った」「収入5000円 副業」\n"
            "📊 **収支確認** → 「今月見せて」「先月の明細」「3月の収支」\n"
            "📂 **カテゴリ別** → 「今月の食費いくら？」「先月の交通費教えて」\n"
            "✏️ **修正** → 記録後に「備考に領収書ありって入れて」「金額800円に直して」\n"
            "🗑️ **削除** → 「消して」「取り消し」\n"
            "✍️ **手入力** → 「手入力」または写真失敗時に自動移行\n"
            "📋 **確定申告** → 「まとめて」「確定申告用ファイル作って」\n"
            "⚙️ **初期設定** → 「設定」で事業内容を再設定\n"
            "🔄 **定期費用** → 「毎月1日に家賃8万円」\n"
            "💰 **予算設定** → 「食費の予算3万円」\n"
            "🔍 **検索** → 「セブンの記録」「5000円以上の支出」\n"
            "📊 **経理モード** → 「経理モード」で税務・経理AIに切り替え"
        )
        return

    # unknown：家計簿と無関係なメッセージはスルー
    await bot.process_commands(message)


if __name__ == "__main__":
    print("kakeibo Bot starting...")
    bot.run(DISCORD_TOKEN)
